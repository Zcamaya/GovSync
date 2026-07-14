from openpyxl import load_workbook

path = r'e:\My Program\06-2026 SSS MEGS - Copy\ALFAMART TRADING PHILS 06-2026.xlsx'
wb = load_workbook(path, data_only=True, read_only=True)
print('sheets', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    print('--- SHEET', name, 'rows', len(rows), '---')
    for row in rows[:10]:
        print(row)
    print()
wb.close()
