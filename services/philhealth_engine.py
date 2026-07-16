import datetime
import os
import re
import sys
import json
import uuid
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QFileDialog, QMessageBox, QComboBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QTabWidget,
    QProgressBar, QFrame, QScrollArea, QGridLayout, QAbstractItemView
)
from PySide6.QtGui import QColor
import openpyxl
from openpyxl.styles import Font as XlFont, PatternFill, Alignment as XlAlignment, Border as XlBorder, Side as XlSide
from models.history import HistoryRecord
from repositories.history_repository import HistoryRepository
from services.auth_manager import database_path, get_active_account
from services.dashboard_service import get_account_username
from shared.ui import set_exit_icon
from constants.styles import AppStyles
from widgets.shared_table import SharedTable


# ---------------------------------------------------------
# UI COMPONENTS
# ---------------------------------------------------------

class MetricCard(QFrame):
    """Interactive dashboard card component."""
    def __init__(self, title, value, color_hex="#3b82f6", parent=None):
        super().__init__(parent)
        self.title = title
        self.value = str(value)
        self.color_hex = color_hex
        self.setFrameShape(QFrame.StyledPanel)
        self.setCursor(Qt.PointingHandCursor)
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(4)
        layout.setContentsMargins(16, 16, 16, 16)
        
        self.title_lbl = QLabel(self.title)
        self.title_lbl.setStyleSheet("color: #94a3b8; font-size: 12px; font-weight: 600; text-transform: uppercase; background: transparent;")
        
        self.value_lbl = QLabel(self.value)
        self.value_lbl.setStyleSheet(f"color: {self.color_hex}; font-size: 28px; font-weight: bold; background: transparent;")
        
        layout.addWidget(self.title_lbl)
        layout.addWidget(self.value_lbl)
        
        self.setStyleSheet(f"""
            MetricCard {{
                background-color: #1e293b;
                border: 1px solid #334155;
                border-radius: 8px;
            }}
            MetricCard:hover {{
                background-color: #334155;
                border-color: #475569;
            }}
        """)

    def update_value(self, new_val):
        self.value = str(new_val)
        self.value_lbl.setText(self.value)


class HistoryCard(QFrame):
    """Card representing a historical processing run."""
    def __init__(self, record_data, delete_cb, view_cb, parent=None):
        super().__init__(parent)
        self.record = record_data
        self.delete_cb = delete_cb
        self.view_cb = view_cb
        self.setCursor(Qt.PointingHandCursor)
        self.setFixedSize(280, 140)
        self.init_ui()

    def init_ui(self):
        self.setStyleSheet("""
            HistoryCard {
                background-color: #1e293b;
                border: 1px solid #334155;
                border-radius: 8px;
            }
            HistoryCard:hover {
                background-color: #334155;
                border-color: #3b82f6;
            }
        """)
        layout = QVBoxLayout(self)
        
        # Upper Layout for X Button & Title
        top_layout = QHBoxLayout()
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(8)
        
        self.del_btn = QPushButton()
        self.del_btn.setFixedSize(22, 22)
        self.del_btn.setCursor(Qt.PointingHandCursor)
        set_exit_icon(self.del_btn, "#fb7185", 12)
        self.del_btn.setStyleSheet("""
            QPushButton {
                background-color: #334155;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #ef4444;
            }
        """)
        self.del_btn.clicked.connect(self.request_delete)
        
        title = QLabel(self.record.get("month_year", "Unknown Date"))
        title.setStyleSheet("font-size: 14px; font-weight: bold; color: #f8fafc; border: none; background: transparent;")
        
        top_layout.addWidget(self.del_btn)
        top_layout.addWidget(title)
        top_layout.addStretch()
        layout.addLayout(top_layout)

        # Metrics Layout
        metrics_layout = QVBoxLayout()
        metrics_layout.setSpacing(2)
        
        def make_row(label, val, color):
            row = QHBoxLayout()
            lbl = QLabel(label)
            lbl.setStyleSheet("color: #94a3b8; font-size: 11px; border: none; background: transparent;")
            v_lbl = QLabel(str(val))
            v_lbl.setStyleSheet(f"color: {color}; font-weight: bold; font-size: 12px; border: none; background: transparent;")
            row.addWidget(lbl)
            row.addStretch()
            row.addWidget(v_lbl)
            return row

        metrics_layout.addLayout(make_row("Total Employees:", self.record.get("total_count", 0), "#3b82f6"))
        metrics_layout.addLayout(make_row("Missing:", self.record.get("missing_count", 0), "#f59e0b"))
        metrics_layout.addLayout(make_row("Newly Hired:", self.record.get("new_count", 0), "#10b981"))
        layout.addLayout(metrics_layout)
        layout.addStretch()

    def request_delete(self):
        self.delete_cb(self.record["id"], self.record["month_year"])

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.view_cb(self.record)


class HistoryDetailPopup(QFrame):
    """Popup showing full tables from a historical run."""
    def __init__(self, record, parent=None):
        super().__init__(parent)
        self.record = record
        self.init_ui()
        
    def init_ui(self):
        self.setMinimumSize(700, 500)
        self.setStyleSheet("""
            HistoryDetailPopup {
                background-color: #0b0f19;
                border: 2px solid #334155;
                border-radius: 12px;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        
        header_layout = QHBoxLayout()
        title_lbl = QLabel(f"Historical Records: {self.record.get('month_year', '')}")
        title_lbl.setStyleSheet("font-size: 16px; font-weight: bold; color: #3b82f6; background: transparent; border: none;")
        
        close_btn = QPushButton()
        close_btn.setFixedSize(28, 28)
        set_exit_icon(close_btn, "#94a3b8", 12)
        close_btn.setStyleSheet("""
            QPushButton { background-color: #1e293b; border: none; border-radius: 14px; }
            QPushButton:hover { background-color: #ef4444; }
        """)
        close_btn.clicked.connect(self.deleteLater)
        
        header_layout.addWidget(title_lbl)
        header_layout.addStretch()
        header_layout.addWidget(close_btn)
        layout.addLayout(header_layout)
        
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(AppStyles.UNIFIED_TAB_STYLE)
        headers = ["Client", "PhilHealth No", "Employee Name", "Birthdate"]
        
        self.tabs.addTab(self.create_table(headers, self.record.get("data_total", [])), "Total Active")
        self.tabs.addTab(self.create_table(headers, self.record.get("data_missing", [])), "Missing")
        self.tabs.addTab(self.create_table(headers, self.record.get("data_new", [])), "Newly Hired")
        
        layout.addWidget(self.tabs)

    def create_table(self, headers, data):
        table = SharedTable(headers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setStyleSheet(AppStyles.TABLE_CANONICAL)
        
        table.setRowCount(len(data))
        for row, items in enumerate(data):
            for col, val in enumerate(items):
                item = QTableWidgetItem(str(val))
                item.setForeground(QColor("#e2e8f0"))
                if col in [0, 1, 3]:
                    item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, col, item)
        return table


class DetailPopup(QFrame):
    """Persistent, closeable detail card overlay for current active metrics."""
    def __init__(self, title, headers, data_list, parent=None):
        super().__init__(parent)
        self.title = title
        self.headers = headers
        self.data_list = data_list
        self.init_ui()
        
    def init_ui(self):
        self.setMinimumSize(600, 450)
        self.setStyleSheet("""
            DetailPopup {
                background-color: #0b0f19;
                border: 2px solid #334155;
                border-radius: 12px;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        
        header_layout = QHBoxLayout()
        title_lbl = QLabel(self.title)
        title_lbl.setStyleSheet("font-size: 15px; font-weight: bold; color: #3b82f6; border: none;")
        
        close_btn = QPushButton()
        close_btn.setFixedSize(28, 28)
        set_exit_icon(close_btn, "#94a3b8", 12)
        close_btn.setStyleSheet("""
            QPushButton { background-color: #1e293b; border: none; border-radius: 14px; }
            QPushButton:hover { background-color: #ef4444; }
        """)
        close_btn.clicked.connect(self.request_close)
        
        header_layout.addWidget(title_lbl)
        header_layout.addStretch()
        header_layout.addWidget(close_btn)
        layout.addLayout(header_layout)
        
        self.table = SharedTable(self.headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setStyleSheet(AppStyles.TABLE_CANONICAL)
        
        self.table.setRowCount(len(self.data_list))
        for row, items in enumerate(self.data_list):
            for col, val in enumerate(items):
                item = QTableWidgetItem(str(val))
                item.setForeground(QColor("#e2e8f0"))
                if col in [0, 1, 3]:
                    item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, col, item)
                
        layout.addWidget(self.table)
        
    def request_close(self):
        confirm = QMessageBox.question(self, "Confirm Dismissal", f"Close the detail view for '{self.title}'?", QMessageBox.Yes | QMessageBox.No)
        if confirm == QMessageBox.Yes:
            self.setParent(None)
            self.deleteLater()


# ---------------------------------------------------------
# MAIN APPLICATION
# ---------------------------------------------------------

class PhicExtractorApp(QWidget):

    def __init__(self):
        super().__init__()
        self.MASTER_HEADERS = [
            "custid", "phealthno", "Monthly Salary", "birthdate", "lastname",
            "firstname", "middlename", "phealthee", "phealther", "ee ded",
            "er ded", "total ded", "phealthrem", "position", "datehired",
            "basicrate", "CLIENT"
        ]
        
        self.cache_total_active = []
        self.cache_missing = []
        self.cache_newly_hired = []
        self.history_records = []
        self.history_repository = HistoryRepository(database_path())
        
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("PHIC Master Data Consolidator & Tracker")
        self.setMinimumWidth(1000)
        self.setMinimumHeight(750)
        
        self.base_layout = QVBoxLayout(self)
        self.base_layout.setContentsMargins(0, 0, 0, 0)
        
        self.central_widget = QWidget()
        self.base_layout.addWidget(self.central_widget)
        
        main_layout = QVBoxLayout(self.central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        self.setStyleSheet("""
            QWidget { background-color: #0f172a; color: #f8fafc; font-family: 'Segoe UI', Arial, sans-serif; }
            QLabel { color: #94a3b8; font-weight: 500; }
            QLineEdit, QComboBox {
                background-color: #1e293b; border: 1px solid #334155; border-radius: 6px; padding: 6px 12px; color: #f8fafc;
            }
            QLineEdit:focus, QComboBox:focus { border: 1px solid #3b82f6; }
            QPushButton {
                background-color: #1e293b; border: 1px solid #334155; border-radius: 6px; padding: 6px 14px; color: #e2e8f0; font-weight: 500;
            }
            QPushButton:hover { background-color: #334155; border-color: #475569; }
            QTabWidget::pane { border: 1px solid #1e293b; background-color: #0b0f19; border-radius: 8px; }
            QTabBar::tab {
                background-color: #111827; color: #94a3b8; padding: 8px 20px; margin-right: 4px;
                border-top-left-radius: 6px; border-top-right-radius: 6px; font-weight: bold;
            }
            QTabBar::tab:selected { background-color: #0b0f19; color: #3b82f6; border-bottom: 2px solid #3b82f6; }
            QScrollBar:vertical { border: none; background: #0f172a; width: 10px; margin: 0px 0px 0px 0px; }
            QScrollBar::handle:vertical { background: #334155; min-height: 20px; border-radius: 14px; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        """)

        # DIRECTORIES
        dir_widget = QWidget()
        dir_widget.setStyleSheet("background-color: #111827; border-radius: 8px; padding: 10px;")
        dir_layout = QVBoxLayout(dir_widget)
        
        src_layout = QHBoxLayout()
        self.src_label = QLabel("Source Directory:")
        self.src_label.setFixedWidth(120)
        self.src_input = QLineEdit()
        self.src_btn = QPushButton("Browse...")
        self.src_btn.clicked.connect(self.browse_source)
        src_layout.addWidget(self.src_label)
        src_layout.addWidget(self.src_input)
        src_layout.addWidget(self.src_btn)
        dir_layout.addLayout(src_layout)

        save_layout = QHBoxLayout()
        self.save_label = QLabel("Save Directory:")
        self.save_label.setFixedWidth(120)
        self.save_input = QLineEdit()
        self.save_input.textChanged.connect(self.load_history) 
        self.save_btn = QPushButton("Browse...")
        self.save_btn.clicked.connect(self.browse_save)
        save_layout.addWidget(self.save_label)
        save_layout.addWidget(self.save_input)
        save_layout.addWidget(self.save_btn)
        dir_layout.addLayout(save_layout)
        main_layout.addWidget(dir_widget)

        # SETTINGS
        name_layout = QHBoxLayout()
        self.fn_label = QLabel("Output Base Name:")
        self.fn_input = QLineEdit("PHIC ALL CLIENTS (MEGS)")

        self.month_combo = QComboBox()
        self.months_list = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(self.months_list)
        self.month_combo.setStyleSheet(AppStyles.GLOBAL_DROPDOWN)

        self.year_combo = QComboBox()
        current_year = datetime.datetime.now().year
        self.year_combo.addItems([str(y) for y in range(current_year - 5, current_year + 3)])
        self.year_combo.setStyleSheet(AppStyles.GLOBAL_DROPDOWN)

        today = datetime.datetime.now()
        last_day_of_prev_month = today.replace(day=1) - datetime.timedelta(days=1)
        self.month_combo.setCurrentIndex(last_day_of_prev_month.month - 1)
        self.year_combo.setCurrentText(str(last_day_of_prev_month.year))

        name_layout.addWidget(self.fn_label)
        name_layout.addWidget(self.fn_input)
        name_layout.addWidget(QLabel("Month:"))
        name_layout.addWidget(self.month_combo)
        name_layout.addWidget(QLabel("Year:"))
        name_layout.addWidget(self.year_combo)
        main_layout.addLayout(name_layout)

        # PROCESS BUTTON
        self.process_btn = QPushButton("Start Processing & Reconciliation")
        self.process_btn.setStyleSheet("""
            QPushButton { background-color: #2563eb; border: none; font-weight: bold; font-size: 13px; padding: 10px; color: white; }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        self.process_btn.clicked.connect(self.process_files)
        main_layout.addWidget(self.process_btn)

        # DASHBOARD METRIC CARDS
        self.cards_layout = QHBoxLayout()
        self.cards_layout.setSpacing(12)
        
        self.card_month = MetricCard("Applicable Month", "Pending", "#10b981")
        self.card_total = MetricCard("Total Active", "0", "#3b82f6")
        self.card_missing = MetricCard("Missing", "0", "#f59e0b")
        self.card_new = MetricCard("New Employees", "0", "#ec4899")
        
        self.card_total.mousePressEvent = lambda e: self.show_current_detail_popup("Total Active Staff", ["Client", "PhilHealth No", "Employee Name", "Birthdate"], self.cache_total_active)
        self.card_missing.mousePressEvent = lambda e: self.show_current_detail_popup("Missing Employees", ["Client", "PhilHealth No", "Employee Name", "Birthdate"], self.cache_missing)
        self.card_new.mousePressEvent = lambda e: self.show_current_detail_popup("Newly Hired Entries", ["Client", "PhilHealth No", "Employee Name", "Birthdate"], self.cache_newly_hired)
        
        self.cards_layout.addWidget(self.card_month)
        self.cards_layout.addWidget(self.card_total)
        self.cards_layout.addWidget(self.card_missing)
        self.cards_layout.addWidget(self.card_new)
        main_layout.addLayout(self.cards_layout)

        # TABS 
        self.tab_container = QTabWidget()
        
        self.missing_table = SharedTable(["Client Handle", "PhilHealth No", "Employee Name", "Birthdate"])
        self.setup_table_headers(self.missing_table, ["Client Handle", "PhilHealth No", "Employee Name", "Birthdate"])
        self.tab_container.addTab(self.missing_table, "Missing Names")

        self.newly_hired_table = SharedTable(["Client Handle", "PhilHealth No", "Employee Name", "Birthdate"])
        self.setup_table_headers(self.newly_hired_table, ["Client Handle", "PhilHealth No", "Employee Name", "Birthdate"])
        self.tab_container.addTab(self.newly_hired_table, "Newly Hired")

        self.summary_table = SharedTable(["Metric Indicator Description", "Headcount Figures"])
        self.setup_table_headers(self.summary_table, ["Metric Indicator Description", "Headcount Figures"])
        self.tab_container.addTab(self.summary_table, "Total Summary Metrics")

        self.history_tab = QWidget()
        history_layout = QVBoxLayout(self.history_tab)
        
        self.history_scroll = QScrollArea()
        self.history_scroll.setWidgetResizable(True)
        self.history_scroll.setStyleSheet("QScrollArea { border: none; background-color: transparent; }")
        
        self.history_grid_widget = QWidget()
        self.history_grid_widget.setStyleSheet("background-color: transparent;")
        self.history_grid_layout = QGridLayout(self.history_grid_widget)
        self.history_grid_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        
        self.history_scroll.setWidget(self.history_grid_widget)
        history_layout.addWidget(self.history_scroll)
        
        self.tab_container.addTab(self.history_tab, "History")
        main_layout.addWidget(self.tab_container)
        
        # LOADING OVERLAY
        self.loading_overlay = QFrame(self)
        self.loading_overlay.setObjectName("Overlay")
        self.loading_overlay.setStyleSheet("QFrame#Overlay { background-color: rgba(15, 23, 42, 0.85); }")
        overlay_layout = QVBoxLayout(self.loading_overlay)
        overlay_layout.setAlignment(Qt.AlignCenter)
        
        self.loading_label = QLabel("Processing Data Records...")
        self.loading_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #f8fafc; margin-bottom: 8px;")
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedWidth(400)
        self.progress_bar.setStyleSheet("""
            QProgressBar { border: 1px solid #334155; border-radius: 6px; background-color: #1e293b; text-align: center; color: white; font-weight: bold; }
            QProgressBar::chunk { background-color: #2563eb; border-radius: 5px; }
        """)
        overlay_layout.addWidget(self.loading_label)
        overlay_layout.addWidget(self.progress_bar)
        self.loading_overlay.hide()

    # ---------------------------------------------------------
    # DATA & HISTORY LOGIC
    # ---------------------------------------------------------
    
    def _load_previous_month_identities_from_history(self, period_label):
        for record in reversed(self.history_records):
            if record.get("month_year") != period_label:
                continue

            identities = set()
            for row in record.get("data_total", []):
                if len(row) < 4:
                    continue
                client, phealth, name, birthdate = row[:4]
                identities.add((str(phealth or "").strip(), str(birthdate or "").strip(), str(name or "").strip(), str(client or "").strip()))
            return identities

        return set()

    def _emit_progress(self, progress_callback, percent, filename="", message=""):
        if progress_callback:
            progress_callback(int(max(0, min(100, percent))), filename, message)

    def load_history(self):
        self.history_records = []
        username = get_account_username()
        try:
            records = self.history_repository.list_by_account(username)
            loaded_records = [
                record.extra.get("payload", {})
                for record in records
                if isinstance(record.extra.get("payload", {}), dict)
            ]
            self.history_records = list(reversed(loaded_records))
        except Exception as e:
            print("Error loading history:", e)
        self.render_history_grid()

    def save_history_record(self, record_data):
        self.history_records.append(record_data)
        username = get_account_username()
        try:
            self.history_repository.save(
                HistoryRecord(
                    id=str(record_data.get("id", "")),
                    account_username=username,
                    module="philhealth",
                    period_label=str(record_data.get("month_year", "")),
                    payload_json=json.dumps(record_data),
                    created_at=record_data.get("created_at"),
                    extra={"payload": dict(record_data)},
                )
            )
            self.render_history_grid()
        except Exception as e:
            QMessageBox.warning(self, "Save Error", f"Could not save history log.\n{str(e)}")

    def delete_history_record(self, record_id, month_year_label):
        confirm = QMessageBox.question(self, "Delete History", f"Are you sure you want to delete the history log for {month_year_label}?", QMessageBox.Yes | QMessageBox.No)
        if confirm == QMessageBox.Yes:
            self.history_records = [r for r in self.history_records if r["id"] != record_id]
            try:
                username = get_account_username()
                self.history_repository.delete_by_id(record_id, username)
                self.render_history_grid()
            except Exception as e:
                print("Error saving after delete:", e)

    def render_history_grid(self):
        while self.history_grid_layout.count():
            child = self.history_grid_layout.takeAt(0)
            if child.widget(): child.widget().deleteLater()
                
        col_count = 3
        for idx, record in enumerate(reversed(self.history_records)):
            row = idx // col_count
            col = idx % col_count
            card = HistoryCard(record, self.delete_history_record, self.show_history_detail_popup)
            self.history_grid_layout.addWidget(card, row, col)

    def show_history_detail_popup(self, record):
        popup = HistoryDetailPopup(record, self)
        popup.move(50, 150)
        popup.show()
        popup.raise_()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.resize(event.size())

    def show_current_detail_popup(self, title, headers, data):
        if not data:
            QMessageBox.information(self, "No Records", f"No metric records found to review for '{title}'.")
            return
        popup = DetailPopup(title, headers, data, self)
        popup.move(60, 220)
        popup.show()
        popup.raise_()

    def setup_table_headers(self, table, headers):
        table.set_columns(headers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setStyleSheet(AppStyles.TABLE_CANONICAL)

    def browse_source(self):
        dir_path = QFileDialog.getExistingDirectory(self, "Select Source Directory")
        if dir_path: self.src_input.setText(dir_path)

    def browse_save(self):
        dir_path = QFileDialog.getExistingDirectory(self, "Select Save Directory")
        if dir_path: self.save_input.setText(dir_path)

    def clean_filename(self, filename):
        name_without_ext = os.path.splitext(filename)[0]
        letters_only = re.sub(r"[^a-zA-Z\s]", "", name_without_ext)
        return re.sub(r"\s+", " ", letters_only).strip()

    def clean_phealth_no(self, val):
        if val is None: return ""
        digits = re.sub(r"\D", "", str(val))
        return digits.zfill(12) if digits else ""

    def parse_to_date_object(self, val):
        if val is None or str(val).strip() == "": return None
        if isinstance(val, (datetime.datetime, datetime.date)): return val
        if isinstance(val, (int, float)):
            try: return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(val))).date()
            except: return None
        
        clean_str = str(val).strip()
        if clean_str.isdigit():
            try:
                return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(clean_str))).date()
            except (ValueError, OverflowError):
                pass

        for pattern in ["%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%b %d, %Y", "%B %d, %Y", "%m/%d/%y", "%d-%b-%y"]:
            try:
                return datetime.datetime.strptime(clean_str, pattern).date()
            except ValueError:
                continue
        return clean_str

    def get_previous_month_info(self, current_month_str, current_year_str):
        try:
            m_idx = self.months_list.index(current_month_str) + 1
            year = int(current_year_str)
            prev_month_date = datetime.date(year, m_idx, 1) - datetime.timedelta(days=1)
            return self.months_list[prev_month_date.month - 1], str(prev_month_date.year)
        except (ValueError, IndexError, TypeError):
            return None, None

    def read_existing_master_identities(self, file_path):
        identities = set()
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if "PHIC Master" in wb.sheetnames:
                ws = wb["PHIC Master"]
                for r_idx in range(2, ws.max_row + 1):
                    phealth = self.clean_phealth_no(ws.cell(row=r_idx, column=2).value)
                    bdate_val = ws.cell(row=r_idx, column=4).value
                    parsed_bdate = self.parse_to_date_object(bdate_val)
                    bdate_str = parsed_bdate.strftime("%m/%d/%Y") if isinstance(parsed_bdate, (datetime.datetime, datetime.date)) else str(bdate_val or "").strip()
                    lname = str(ws.cell(row=r_idx, column=5).value or "").strip().upper()
                    fname = str(ws.cell(row=r_idx, column=6).value or "").strip().upper()
                    client = str(ws.cell(row=r_idx, column=17).value or "").strip()
                    if phealth or lname or fname:
                        identities.add((phealth, bdate_str, f"{lname}, {fname}", client))
            wb.close()
        except Exception as exc:
            print(f"Error reading existing master identities: {exc}")
        return identities

    def set_ui_enabled(self, enabled: bool):
        self.central_widget.setEnabled(enabled)
        if not enabled:
            self.progress_bar.setRange(0, 0)
            self.loading_overlay.show()
            self.loading_overlay.raise_()
        else:
            self.loading_overlay.hide()

    def generate_match_key(self, phealth, full_name):
        """Creates a highly strict string ID for an employee immune to date formatting differences."""
        ph = str(phealth).strip()
        # Strip absolutely ALL whitespace from the name to prevent typo-based false flags
        name_normalized = re.sub(r'\s+', '', str(full_name).upper())
        return f"{ph}_{name_normalized}"

    def populate_ui_table(self, table_widget, grouped_data):
        table_widget.setRowCount(0)
        row_idx = 0
        # Iterate over sorted client names
        for client in sorted(grouped_data.keys(), key=lambda x: str(x).upper()):
            emps = grouped_data[client]
            for ph, bd, name in emps:
                table_widget.insertRow(row_idx)
                c_item = QTableWidgetItem(str(client))
                ph_item = QTableWidgetItem(str(ph))
                n_item = QTableWidgetItem(str(name))
                b_item = QTableWidgetItem(str(bd))
                
                ph_item.setTextAlignment(Qt.AlignCenter)
                b_item.setTextAlignment(Qt.AlignCenter)

                for item in [c_item, ph_item, n_item, b_item]:
                    item.setForeground(QColor("#e2e8f0"))
                
                table_widget.setItem(row_idx, 0, c_item)
                table_widget.setItem(row_idx, 1, ph_item)
                table_widget.setItem(row_idx, 2, n_item)
                table_widget.setItem(row_idx, 3, b_item)
                row_idx += 1

    def write_custom_variance_sheet(self, wb, sheet_title, grouped_data, header_label, section_title):
        ws = wb.create_sheet(title=sheet_title)
        title_font = XlFont(name="Arial", size=14, bold=True, color="1B5E20")
        client_header_font = XlFont(name="Arial", size=11, bold=True, color="333333")
        table_header_font = XlFont(name="Arial", size=10, bold=True, color="FFFFFF")
        data_font = XlFont(name="Arial", size=10)
        accent_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        thin_border = XlBorder(left=XlSide(style="thin", color="CCCCCC"), right=XlSide(style="thin", color="CCCCCC"), top=XlSide(style="thin", color="CCCCCC"), bottom=XlSide(style="thin", color="CCCCCC"))

        ws.cell(row=2, column=2, value=section_title).font = title_font
        ws.cell(row=3, column=2, value=header_label).font = XlFont(name="Arial", size=10, italic=True, color="666666")

        current_row = 5
        if not grouped_data:
            ws.cell(row=current_row, column=2, value="No variance modifications tracked.").font = XlFont(name="Arial", size=11, italic=True)
            return

        for client in sorted(grouped_data.keys(), key=lambda x: str(x).upper()):
            emps = grouped_data[client]
            ws.cell(row=current_row, column=2, value=client.upper()).font = client_header_font
            current_row += 1
            headers = ["PhilHealth ID", "Employee Name", "Birthdate"]
            for col_idx, text in enumerate(headers, start=2):
                cell = ws.cell(row=current_row, column=col_idx, value=text)
                cell.font = table_header_font; cell.fill = accent_fill; cell.alignment = XlAlignment(horizontal="center")
            current_row += 1

            for ph, bd, name in emps:
                c1 = ws.cell(row=current_row, column=2, value=ph)
                c2 = ws.cell(row=current_row, column=3, value=name)
                c3 = ws.cell(row=current_row, column=4, value=bd)
                c1.alignment = XlAlignment(horizontal="center"); c1.number_format = "@"
                c3.alignment = XlAlignment(horizontal="center")
                for cell in [c1, c2, c3]:
                    cell.font = data_font; cell.border = thin_border
                current_row += 1
            current_row += 2

        for col in range(2, 5):
            max_len = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(max_len + 4, 15)

    def write_summary_sheet(self, wb, total_curr, total_prev, missing, added):
        ws = wb.create_sheet(title="Total Employees Summary")
        title_font = XlFont(name="Arial", size=14, bold=True, color="0D47A1")
        header_font = XlFont(name="Arial", size=11, bold=True, color="FFFFFF")
        metric_font = XlFont(name="Arial", size=11, bold=True)
        value_font = XlFont(name="Arial", size=11, color="333333")
        blue_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        thin_border = XlBorder(left=XlSide(style="thin", color="CCCCCC"), right=XlSide(style="thin", color="CCCCCC"), top=XlSide(style="thin", color="CCCCCC"), bottom=XlSide(style="thin", color="CCCCCC"))

        ws.cell(row=2, column=2, value="PHIC MONTHLY RECONCILIATION SUMMARY").font = title_font
        for col_idx, text in enumerate(["Metric Description", "Headcount Value"], start=2):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = header_font; cell.fill = blue_fill; cell.alignment = XlAlignment(horizontal="center")

        metrics = [
            ("Total Active Staffs (Current Month)", total_curr),
            ("Total Active Staffs (Previous Month Baseline)", total_prev),
            ("Total Missing Headcounts", missing),
            ("Total Newly Hired Headcounts", added),
        ]

        self.summary_table.setRowCount(0)
        for idx, (label, val) in enumerate(metrics, start=5):
            c1 = ws.cell(row=idx, column=2, value=label)
            c2 = ws.cell(row=idx, column=3, value=val)
            c1.font = metric_font; c2.font = value_font
            c2.alignment = XlAlignment(horizontal="right")
            c1.border = thin_border; c2.border = thin_border

            ui_row = idx - 5
            self.summary_table.insertRow(ui_row)
            item_lbl = QTableWidgetItem(label)
            item_val = QTableWidgetItem(str(val))
            item_val.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item_lbl.setForeground(QColor("#e2e8f0"))
            item_val.setForeground(QColor("#3b82f6"))
            self.summary_table.setItem(ui_row, 0, item_lbl)
            self.summary_table.setItem(ui_row, 1, item_val)

        ws.column_dimensions["B"].width = 45; ws.column_dimensions["C"].width = 20

    def process_files(self, progress_callback=None):
        source_dir = self.src_input.text().strip()
        save_dir = self.save_input.text().strip()
        base_filename = self.fn_input.text().strip()
        selected_month = self.month_combo.currentText()
        selected_year = self.year_combo.currentText()

        if not source_dir or not os.path.exists(source_dir) or not save_dir or not base_filename:
            QMessageBox.warning(self, "Error", "Ensure all operational parameters are correct.")
            return

        if not os.path.exists(save_dir): os.makedirs(save_dir)
        self.set_ui_enabled(False)
        QApplication.processEvents()
        self._emit_progress(progress_callback, 0, "", "Preparing PhilHealth reconciliation...")

        prev_month, prev_year = self.get_previous_month_info(selected_month, selected_year)
        prev_period_label = f"{prev_month} {prev_year}" if prev_month and prev_year else ""
        previous_month_set = self._load_previous_month_identities_from_history(prev_period_label) if prev_period_label else set()

        if not previous_month_set and prev_month and prev_year:
            prev_filename = f"{base_filename} {prev_month} {prev_year}.xlsx"
            prev_filepath = os.path.join(save_dir, prev_filename)
            previous_month_set = self.read_existing_master_identities(prev_filepath) if os.path.exists(prev_filepath) else set()

        excel_files = [f for f in os.listdir(source_dir) if f.endswith((".xlsx", ".xls", ".xlsm"))]
        if not excel_files:
            self.set_ui_enabled(True)
            self._emit_progress(progress_callback, 100, "", "No source files found.")
            return

        file_plan = []
        total_work_units = 0
        for file in excel_files:
            file_path = os.path.join(source_dir, file)
            row_estimate = 1
            try:
                probe_wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                if "PHIC" in probe_wb.sheetnames:
                    probe_ws = probe_wb["PHIC"]
                    row_estimate = max(1, probe_ws.max_row - 1)
                probe_wb.close()
            except Exception:
                row_estimate = 1
            file_plan.append((file, file_path, row_estimate))
            total_work_units += row_estimate
        total_work_units = max(total_work_units, 1)

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "PHIC All CLient"
        out_ws.append(self.MASTER_HEADERS)
        
        green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = XlFont(name="Arial", size=11, bold=True, color="FFFFFF")
        for col_idx in range(1, len(self.MASTER_HEADERS) + 1):
            cell = out_ws.cell(row=1, column=col_idx)
            cell.fill = green_fill; cell.font = header_font
        out_ws.freeze_panes = "A2"

        red_font = XlFont(name="Arial", size=10, color="FF0000")
        normal_font = XlFont(name="Arial", size=10)
        currency_format = "#,##0.00"; date_format = "mm/dd/yyyy"

        output_current_row = 2
        current_month_set = set()
        completed_work_units = 0

        try:
            for file_index, (file, file_path, row_estimate) in enumerate(file_plan, start=1):
                self._emit_progress(
                    progress_callback,
                    int((completed_work_units / total_work_units) * 100),
                    file,
                    f"Scanning file {file_index} of {len(file_plan)}",
                )
                QApplication.processEvents()
                try:
                    in_wb = openpyxl.load_workbook(file_path, data_only=True)
                    if "PHIC" in in_wb.sheetnames:
                        in_ws = in_wb["PHIC"]
                        client_name = self.clean_filename(file)
                        file_rows_done = 0

                        for row_idx in range(2, in_ws.max_row + 1):
                            if row_idx % 50 == 0:
                                QApplication.processEvents()
                            file_rows_done = min(row_idx - 1, row_estimate)
                            current_units = completed_work_units + file_rows_done
                            self._emit_progress(
                                progress_callback,
                                int((current_units / total_work_units) * 100),
                                file,
                                f"Reading row {row_idx - 1} in {file}",
                            )
                            raw_phealth = in_ws.cell(row=row_idx, column=2).value
                            phealth_val = self.clean_phealth_no(raw_phealth)
                            if not phealth_val: break

                            custid = in_ws.cell(row=row_idx, column=1).value
                            birthdate_raw = in_ws.cell(row=row_idx, column=4).value
                            birthdate_obj = self.parse_to_date_object(birthdate_raw)
                            birthdate_display_str = birthdate_obj.strftime("%m/%d/%Y") if isinstance(birthdate_obj, (datetime.date, datetime.datetime)) else str(birthdate_raw or "").strip()

                            lastname = in_ws.cell(row=row_idx, column=5).value
                            firstname = in_ws.cell(row=row_idx, column=6).value
                            middlename = in_ws.cell(row=row_idx, column=7).value
                            phealthee = in_ws.cell(row=row_idx, column=8).value
                            phealther = in_ws.cell(row=row_idx, column=9).value
                            position = in_ws.cell(row=row_idx, column=14).value
                            datehired_raw = in_ws.cell(row=row_idx, column=15).value
                            datehired_obj = self.parse_to_date_object(datehired_raw)
                            basicrate = in_ws.cell(row=row_idx, column=16).value

                            full_name_str = f"{str(lastname or '').strip().upper()}, {str(firstname or '').strip().upper()}"
                            current_month_set.add((phealth_val, birthdate_display_str, full_name_str, client_name))

                            r = output_current_row
                            out_ws.cell(row=r, column=1, value=custid).font = normal_font
                            phealth_cell = out_ws.cell(row=r, column=2, value=phealth_val)
                            phealth_cell.number_format = "@"; phealth_cell.font = normal_font

                            out_ws.cell(row=r, column=3, value=f"=M{r}/0.05").number_format = currency_format
                            out_ws.cell(row=r, column=3).font = normal_font

                            cell_bd = out_ws.cell(row=r, column=4, value=birthdate_obj)
                            cell_bd.number_format = date_format; cell_bd.font = normal_font

                            out_ws.cell(row=r, column=5, value=lastname).font = normal_font
                            out_ws.cell(row=r, column=6, value=firstname).font = normal_font
                            out_ws.cell(row=r, column=7, value=middlename).font = normal_font
                            out_ws.cell(row=r, column=8, value=phealthee).number_format = currency_format
                            out_ws.cell(row=r, column=9, value=phealther).number_format = currency_format
                            out_ws.cell(row=r, column=8).font = normal_font; out_ws.cell(row=r, column=9).font = normal_font

                            out_ws.cell(row=r, column=10, value=f'=IF(H{r}<250, 250-H{r}, "")').number_format = currency_format
                            out_ws.cell(row=r, column=11, value=f'=IF(I{r}<250, 250-I{r}, "")').number_format = currency_format
                            out_ws.cell(row=r, column=12, value=f"=SUM(J{r}:K{r})").number_format = currency_format
                            out_ws.cell(row=r, column=13, value=f"=SUM(H{r}:K{r})").number_format = currency_format
                            out_ws.cell(row=r, column=10).font = red_font; out_ws.cell(row=r, column=11).font = red_font
                            out_ws.cell(row=r, column=12).font = red_font; out_ws.cell(row=r, column=13).font = normal_font

                            out_ws.cell(row=r, column=14, value=position).font = normal_font
                            cell_dh = out_ws.cell(row=r, column=15, value=datehired_obj)
                            cell_dh.number_format = date_format; cell_dh.font = normal_font
                            out_ws.cell(row=r, column=16, value=basicrate).number_format = currency_format
                            out_ws.cell(row=r, column=16).font = normal_font
                            out_ws.cell(row=r, column=17, value=client_name).font = normal_font

                            output_current_row += 1
                    in_wb.close()
                except Exception as e: print(f"Error parsing file {file}: {str(e)}")
                completed_work_units += row_estimate
                self._emit_progress(
                    progress_callback,
                    int((completed_work_units / total_work_units) * 100),
                    file,
                    f"Finished file {file_index} of {len(file_plan)}",
                )

            if output_current_row > 2:
                # Use robust lookup keys immune to birthdate formatting
                prev_lookup = {self.generate_match_key(item[0], item[2]) for item in previous_month_set}
                curr_lookup = {self.generate_match_key(item[0], item[2]) for item in current_month_set}

                # Generate sorted active list: Sort by Client (index 3) -> Name (index 2)
                current_month_list = sorted(list(current_month_set), key=lambda x: (str(x[3]).upper(), str(x[2]).upper()))
                
                # Generate missing/added arrays natively sorted by A-Z
                missing_raw = sorted([i for i in previous_month_set if self.generate_match_key(i[0], i[2]) not in curr_lookup], key=lambda x: (str(x[3]).upper(), str(x[2]).upper()))
                added_raw = sorted([i for i in current_month_list if self.generate_match_key(i[0], i[2]) not in prev_lookup], key=lambda x: (str(x[3]).upper(), str(x[2]).upper()))

                # Update caches formatted for grid consumption [Client, PhilHealth No, Name, Birthdate]
                self.cache_total_active = [[item[3], item[0], item[2], item[1]] for item in current_month_list]
                self.cache_missing = [[item[3], item[0], item[2], item[1]] for item in missing_raw]
                self.cache_newly_hired = [[item[3], item[0], item[2], item[1]] for item in added_raw]

                self.card_month.update_value(f"{selected_month} {selected_year}")
                self.card_total.update_value(len(current_month_list))
                self.card_missing.update_value(len(missing_raw))
                self.card_new.update_value(len(added_raw))

                missing_by_client = {}
                for ph, bd, name, client in missing_raw: missing_by_client.setdefault(client, []).append((ph, bd, name))
                added_by_client = {}
                for ph, bd, name, client in added_raw: added_by_client.setdefault(client, []).append((ph, bd, name))

                self.populate_ui_table(self.missing_table, missing_by_client)
                self.populate_ui_table(self.newly_hired_table, added_by_client)

                self.write_custom_variance_sheet(out_wb, "Missing Employees", missing_by_client, "Omitted staff profiles.", "MISSING RECORD LABELS")
                self.write_custom_variance_sheet(out_wb, "Newly Hired", added_by_client, "Added staff profiles.", "NEW RECRUIT LABELS")
                self.write_summary_sheet(out_wb, len(current_month_list), len(previous_month_set), len(missing_raw), len(added_raw))

                final_filename = f"{base_filename} {selected_month} {selected_year}.xlsx"
                out_wb.save(os.path.join(save_dir, final_filename))
                
                # Update visual JSON cache
                new_history_record = {
                    "id": str(uuid.uuid4()),
                    "month_year": f"{selected_month} {selected_year}",
                    "total_count": len(current_month_list),
                    "previous_count": len(previous_month_set),
                    "missing_count": len(missing_raw),
                    "new_count": len(added_raw),
                    "data_total": self.cache_total_active,
                    "data_missing": self.cache_missing,
                    "data_new": self.cache_newly_hired
                }
                self.save_history_record(new_history_record)

                QMessageBox.information(self, "Success", f"Workbook Compiled Successfully!\n\nReview UI dashboard metrics.")
            else:
                QMessageBox.information(self, "Finished", "No compiled records structured.")
            self._emit_progress(progress_callback, 100, "", "PhilHealth reconciliation finished.")
        finally:
            out_wb.close()
            self.set_ui_enabled(True)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PhicExtractorApp()
    window.show()
    sys.exit(app.exec())
