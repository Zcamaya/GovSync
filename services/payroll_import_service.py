import os
import uuid
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from models.payroll_import import PayrollImport, PayrollImportPlan
from repositories.payroll_repository import PayrollRepository
from services.auth_manager import get_active_account


REQUIRED_EARNINGS_COLUMNS = [
    "batchid",
    "fdate",
    "tdate",
    "custid",
    "company",
    "emplid",
    "lastname",
    "firstname",
    "middlename",
    "suffix",
    "birthdate",
    "basicrate",
    "billrate",
    "sssno",
    "sss",
    "eeshare",
    "ershare",
    "sssrem",
    "er",
    "pagibigno",
    "pagibig",
    "pagibigrem",
    "phealthno",
    "phealth",
    "phealthrem",
    "sssloan",
    "pbigloan",
    "ctr",
    "tinno",
    "datehired",
    "companyid",
    "position",
]


class PayrollImportService:
    def __init__(self, repository: PayrollRepository | None = None):
        self.repository = repository or PayrollRepository()
        self.repository.initialize_schema()

    def inspect_workbook(self, file_path: str) -> tuple[pd.DataFrame, str]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        filename = os.path.basename(file_path)
        if filename.startswith("~$"):
            raise ValueError("Skipped temporary Excel file.")

        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True)
        except PermissionError as exc:
            raise PermissionError(f"Excel file is locked or already open: {filename}") from exc
        except Exception as exc:
            raise ValueError(f"Unable to read workbook {filename}: {exc}") from exc

        try:
            sheet_map = {sheet.lower().strip(): sheet for sheet in workbook.sheetnames}
            if "earnings" not in sheet_map:
                raise ValueError("The workbook must contain an Earnings worksheet.")

            sheet_name = sheet_map["earnings"]
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            raise ValueError("The Earnings worksheet is empty.")

        header_index = None
        for index, row in enumerate(rows):
            values = [self._clean_cell(value) for value in row]
            if any(values):
                header_index = index
                break

        if header_index is None:
            raise ValueError("The Earnings worksheet does not contain headers.")

        headers = [self._normalize_header(value, index) for index, value in enumerate(rows[header_index])]
        data_rows: list[list[Any]] = []
        for row in rows[header_index + 1 :]:
            values = [self._clean_cell(value) for value in row]
            if not any(values):
                continue
            if len(values) < len(headers):
                values.extend([""] * (len(headers) - len(values)))
            elif len(values) > len(headers):
                values = values[: len(headers)]
            data_rows.append(values)

        if not data_rows:
            frame = pd.DataFrame(columns=headers)
        else:
            frame = pd.DataFrame(data_rows, columns=headers)

        frame.columns = [str(column).strip() for column in frame.columns]
        missing_columns = [column for column in REQUIRED_EARNINGS_COLUMNS if column not in frame.columns]
        if missing_columns:
            raise ValueError(
                "The Earnings worksheet is missing required columns: " + ", ".join(missing_columns)
            )
        return frame, sheet_name

    def build_import_plan(
        self,
        file_path: str,
        applicable_month: str,
        progress_callback=None,
    ) -> PayrollImportPlan:
        account = get_active_account() or {}
        employer_id = str(account.get("employer_id") or account.get("employer_name") or account.get("username") or "default")
        frame, _ = self.inspect_workbook(file_path)
        import_id = str(uuid.uuid4())
        file_hash = self.repository.file_hash(file_path)
        from_date = self._coerce_value(frame["fdate"].dropna().iloc[0]) if not frame.empty else ""
        to_date = self._coerce_value(frame["tdate"].dropna().iloc[0]) if not frame.empty else ""
        client_id = self._coerce_value(frame["custid"].dropna().iloc[0]) if not frame.empty else ""
        client_name = self._coerce_value(frame["company"].dropna().iloc[0]) if not frame.empty else ""
        employee_count = int(len(frame))

        existing_import = self.repository.get_import_by_hash(file_hash)
        existing_payroll = self.repository.get_payroll_import(
            employer_id=employer_id,
            applicable_month=applicable_month,
            from_date=from_date,
            to_date=to_date,
            client_id=client_id,
        )
        if existing_payroll is None:
            existing_payroll = self.repository.get_payroll_import_for_month(
                employer_id=employer_id,
                applicable_month=applicable_month,
                client_id=client_id,
            )

        return PayrollImportPlan(
            import_id=import_id,
            file_hash=file_hash,
            employer_id=employer_id,
            applicable_month=applicable_month,
            from_date=from_date,
            to_date=to_date,
            client_id=client_id,
            client_name=client_name,
            employee_count=employee_count,
            original_filename=os.path.basename(file_path),
            records=self._normalize_records(frame, import_id),
            existing_import=existing_import,
            existing_payroll=existing_payroll,
        )

    def persist_import(self, plan: PayrollImportPlan) -> PayrollImport:
        import_record = PayrollImport(
            id=plan.import_id,
            employer_id=plan.employer_id,
            applicable_month=plan.applicable_month,
            from_date=plan.from_date,
            to_date=plan.to_date,
            client_id=plan.client_id,
            client_name=plan.client_name,
            employee_count=plan.employee_count,
            original_filename=plan.original_filename,
            file_hash=plan.file_hash,
            created_at=datetime.utcnow().isoformat(),
            updated_at=datetime.utcnow().isoformat(),
        )
        if plan.existing_payroll is not None:
            existing_import_id = plan.existing_payroll.get("id")
            if existing_import_id:
                self.repository.delete_import(str(existing_import_id))
        self.repository.replace_import(import_record, plan.records)
        return import_record

    def _normalize_records(self, frame: pd.DataFrame, import_id: str) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for index, row in frame.iterrows():
            payload = {"id": f"{import_id}-{index}", "import_id": import_id}
            for column in REQUIRED_EARNINGS_COLUMNS:
                payload[column] = self._coerce_value(row.get(column, ""))
            records.append(payload)
        return records

    def _coerce_value(self, value: Any) -> str:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        return str(value)

    def _clean_cell(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and pd.isna(value):
            return ""
        return str(value).strip()

    def _normalize_header(self, value: Any, index: int) -> str:
        header = self._clean_cell(value).lower()
        if not header:
            header = f"column_{index + 1}"
        return header
