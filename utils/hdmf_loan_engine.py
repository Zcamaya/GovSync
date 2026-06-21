import re

import pandas as pd
from openpyxl import load_workbook

from utils.payroll_engine import safe_save_workbook


def separate_hdmf_loans(earnings_file, monitoring_file):
    if not earnings_file or not monitoring_file:
        raise ValueError("Please select both earnings and monitoring files.")

    monitoring_df = pd.read_excel(monitoring_file, header=6, engine="openpyxl")
    required_columns = {"HDMF", "LOAN TYPE"}
    missing_columns = required_columns.difference(monitoring_df.columns)
    if missing_columns:
        missing = ", ".join(sorted(missing_columns))
        raise ValueError(f"Monitoring file is missing required column(s): {missing}")

    monitoring_df["HDMF_CLEAN"] = (
        monitoring_df["HDMF"].astype(str).str.replace(r"\.0$", "", regex=True)
    )
    monitoring_df["LOAN_TYPE_CLEAN"] = (
        monitoring_df["LOAN TYPE"].astype(str).str.upper().str.strip()
    )

    workbook = load_workbook(earnings_file, keep_vba=earnings_file.lower().endswith(".xlsm"))
    if "HDMF_LOANS" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("HDMF_LOANS sheet not found in the earnings file.")

    worksheet = workbook["HDMF_LOANS"]
    headers = {str(cell.value).strip(): cell.column for cell in worksheet[1] if cell.value}
    pbig_col = headers.get("pagibigno")
    sl_col = headers.get("SL Deduction")
    cl_col = headers.get("CL Deduction")

    if not all([pbig_col, sl_col, cl_col]):
        workbook.close()
        raise ValueError(
            "Required headers missing in HDMF_LOANS row 1: "
            "pagibigno, SL Deduction, CL Deduction."
        )

    updates_sl = apply_loan_type(
        worksheet,
        monitoring_df,
        "MPL",
        sl_col,
        pbig_col,
    )
    updates_cl = apply_loan_type(
        worksheet,
        monitoring_df,
        "CAL",
        cl_col,
        pbig_col,
    )

    safe_save_workbook(workbook, earnings_file)
    return updates_sl, updates_cl


def apply_loan_type(worksheet, monitoring_df, target_loan_type, target_col_idx, pbig_col_idx):
    updates_count = 0

    for row in range(2, worksheet.max_row + 1):
        first_value = str(worksheet.cell(row=row, column=1).value).strip()
        if first_value.upper() == "TOTAL":
            break

        pbig_value = str(worksheet.cell(row=row, column=pbig_col_idx).value).strip()
        pbig_value = re.sub(r"\.0$", "", pbig_value)

        if (not first_value or first_value == "None") and (
            not pbig_value or pbig_value == "None"
        ):
            break
        if not pbig_value or pbig_value == "None":
            continue

        match_records = monitoring_df[
            (monitoring_df["HDMF_CLEAN"] == pbig_value)
            & (monitoring_df["LOAN_TYPE_CLEAN"] == target_loan_type)
        ]

        if match_records.empty:
            continue

        amount_columns = [
            column for column in monitoring_df.columns if "AMOUNT" in str(column).upper()
        ]
        if len(amount_columns) < 2:
            continue

        value = match_records.iloc[0][amount_columns[-2]]
        if pd.notna(value) and str(value).strip() != "":
            worksheet.cell(row=row, column=target_col_idx).value = value
            updates_count += 1

    return updates_count
