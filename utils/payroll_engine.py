import os
import glob
import tempfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from tkinter import Button, END, Label, Listbox, Tk, filedialog


def select_input():
    """Opens a window to select multiple files and/or folders."""
    selected_paths = []

    def on_add_files():
        files = filedialog.askopenfilenames(
            title="Select Excel File(s)",
            filetypes=[("Excel files", "*.xlsx")],
        )
        for file_path in files:
            if file_path not in selected_paths:
                selected_paths.append(file_path)
                listbox.insert(END, os.path.basename(file_path))

    def on_add_folder():
        folder = filedialog.askdirectory(title="Select Folder")
        if folder:
            files = glob.glob(os.path.join(folder, "*.xlsx"))
            for file_path in files:
                if file_path not in selected_paths:
                    selected_paths.append(file_path)
                    listbox.insert(END, os.path.basename(file_path))

    def on_remove():
        selection = listbox.curselection()
        for index in reversed(selection):
            listbox.delete(index)
            del selected_paths[index]

    def on_done():
        root.destroy()

    root = Tk()
    root.title("Select Payroll Files & Folders")
    root.geometry("400x420")

    Label(root, text="Select files/folders to process:", pady=10).pack()

    listbox = Listbox(root, width=50, height=10, selectmode="extended")
    listbox.pack(pady=5)

    Button(root, text="Add File(s)", command=on_add_files, width=25).pack(pady=2)
    Button(root, text="Add Folder", command=on_add_folder, width=25).pack(pady=2)
    Button(root, text="Remove Selected", command=on_remove, width=25, bg="salmon").pack(
        pady=2
    )
    Button(
        root,
        text="Done / Start Processing",
        command=on_done,
        width=25,
        bg="lightgreen",
    ).pack(pady=10)

    root.mainloop()
    return selected_paths


def run_payroll_task(file_path):
    """Main integration function used by the PySide UI."""
    filename = os.path.basename(file_path)
    if filename.startswith("~$"):
        return True, f"Skipped temporary file {filename}", 0

    with pd.ExcelFile(file_path) as workbook:
        sheet_map = {sheet.lower().strip(): sheet for sheet in workbook.sheet_names}

        if "earnings" not in sheet_map:
            raise ValueError(f"No 'Earnings' sheet found in {filename}")

        df_master = pd.read_excel(workbook, sheet_name=sheet_map["earnings"])

    df_master.columns = df_master.columns.str.strip()

    sheets_data = generate_slices(df_master)
    hdmf_count = len(sheets_data.get("HDMF", []))
    save_and_format_sheets(file_path, sheets_data)
    return True, f"Processed and formatted {filename}", hdmf_count


def process_excel_files(file_list):
    if not file_list:
        print("No files selected. Exiting.")
        return

    for file_path in file_list:
        filename = os.path.basename(file_path)
        if filename.startswith("~$"):
            continue

        print(f"\nProcessing file: {filename}")
        try:
            run_payroll_task(file_path)
            print(f"-> Successfully automated and formatted {filename}!")
        except Exception as exc:
            print(f"-> Error processing file {filename}: {exc}")


def generate_slices(df):
    processed_sheets = {}

    phic_cols = [
        "custid",
        "phealthno",
        "birthdate",
        "lastname",
        "firstname",
        "middlename",
        "phealth",
        "tinno",
        "position",
        "datehired",
        "basicrate",
    ]
    df_phic = df[[column for column in phic_cols if column in df.columns]].copy()
    processed_sheets["PHIC"] = df_phic

    hdmf_cols = [
        "pagibigno",
        "companyid",
        "lastname",
        "firstname",
        "middlename",
        "pagibig",
        "tinno",
        "birthdate",
    ]
    processed_sheets["HDMF"] = df[
        [column for column in hdmf_cols if column in df.columns]
    ].copy()

    sss_cols = [
        "sssno",
        "lastname",
        "firstname",
        "middlename",
        "birthdate",
        "sss",
        "eeshare",
        "ershare",
        "sssrem",
        "er",
        "datehired",
        "position",
    ]
    df_sss = df[[column for column in sss_cols if column in df.columns]].copy()
    processed_sheets["SSS"] = df_sss

    hdmf_loan_cols = [
        "custid",
        "pagibigno",
        "companyid",
        "lastname",
        "firstname",
        "middlename",
        "pbigloan",
        "birthdate",
        "tinno",
    ]
    df_hdmf_loans = df[
        [column for column in hdmf_loan_cols if column in df.columns]
    ].copy()
    processed_sheets["HDMF_LOANS"] = df_hdmf_loans

    sss_loan_cols = ["custid", "sssno", "lastname", "firstname", "middlename", "sssloan"]
    df_sss_loans = df[[column for column in sss_loan_cols if column in df.columns]].copy()
    processed_sheets["SSS_LOANS"] = df_sss_loans

    return processed_sheets


def save_and_format_sheets(file_path, sheets_data):
    workbook = load_workbook(file_path, keep_vba=file_path.lower().endswith(".xlsm"))

    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_red = Font(name="Calibri", size=11, color="FF0000")

    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    green_fill = PatternFill(start_color="77D755", end_color="77D755", fill_type="solid")

    thin_side = Side(border_style="thin", color="000000")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total_label = Border(top=Side(border_style="thin", color="000000"))
    border_total_value = Border(
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="double", color="000000"),
    )

    min_contribution = 250
    phealth_percentage = 0.05

    for sheet_name, dataframe in sheets_data.items():
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

        worksheet = workbook.create_sheet(title=sheet_name)
        headers = headers_for_sheet(sheet_name)

        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = font_header
            cell.fill = header_fill
            cell.border = border_all

        for row_index, src_row in enumerate(dataframe.to_dict(orient="records"), 2):
            row_cells = []
            for header in headers:
                row_cells.append(
                    cell_output_value(
                        sheet_name,
                        header,
                        row_index,
                        src_row,
                        min_contribution,
                        phealth_percentage,
                    )
                )

            worksheet.append(row_cells)
            current_row = worksheet[worksheet.max_row]
            apply_row_styles(
                sheet_name,
                src_row,
                headers,
                current_row,
                font_regular,
                font_red,
                green_fill,
                border_all,
            )

        append_totals(
            worksheet,
            sheet_name,
            headers,
            font_bold,
            border_total_label,
            border_total_value,
        )
        autofit_columns(worksheet)

    safe_save_workbook(workbook, file_path)


def safe_save_workbook(workbook, file_path):
    directory = os.path.dirname(os.path.abspath(file_path))
    suffix = os.path.splitext(file_path)[1] or ".xlsx"
    temp_path = ""

    try:
        with tempfile.NamedTemporaryFile(
            delete=False,
            dir=directory,
            prefix="~govsync-save-",
            suffix=suffix,
        ) as temp_file:
            temp_path = temp_file.name

        workbook.save(temp_path)
        load_workbook(temp_path, keep_vba=file_path.lower().endswith(".xlsm")).close()
        os.replace(temp_path, file_path)
    except Exception:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        raise
    finally:
        workbook.close()


def headers_for_sheet(sheet_name):
    headers = {
        "PHIC": [
            "custid",
            "phealthno",
            "Monthly Salary",
            "birthdate",
            "lastname",
            "firstname",
            "middlename",
            "phealthee",
            "phealther",
            "ee ded",
            "er ded",
            "total ded",
            "phealthrem",
            "position",
            "datehired",
            "basicrate",
        ],
        "HDMF": [
            "pagibigno",
            "companyid",
            "lastname",
            "firstname",
            "middlename",
            "ee",
            "er",
            "tinno",
            "birthdate",
        ],
        "SSS": [
            "sssno",
            "lastname",
            "firstname",
            "middlename",
            "Monthly Salary",
            "birthdate",
            "sss",
            "ee",
            "ee ded",
            "ershare",
            "sssrem",
            "er",
            "totalsssrem",
            "datehired",
            "position",
        ],
        "HDMF_LOANS": [
            "custid",
            "pagibigno",
            "companyid",
            "lastname",
            "firstname",
            "middlename",
            "pbigloan",
            "SL Deduction",
            "CL Deduction",
            "Total Deduction",
            "Over Amount Ded",
            "tinno",
            "birthdate",
        ],
        "SSS_LOANS": [
            "custid",
            "sssno",
            "lastname",
            "firstname",
            "middlename",
            "sssloan",
            "SL Deduction",
            "CL Deduction",
            "Total Deduction",
            "Over Amount Ded",
        ],
    }
    return headers[sheet_name]


def cell_output_value(
    sheet_name,
    header,
    row_index,
    src_row,
    min_contribution,
    phealth_percentage,
):
    if sheet_name == "PHIC":
        phealth_value = numeric(src_row.get("phealth", 0))
        ee_deduction = min_contribution - phealth_value
        er_deduction = min_contribution - phealth_value

        if header == "phealthno":
            return padded_number(src_row.get("phealthno", ""), 12)
        if header == "phealthee":
            return src_row.get("phealth", 0)
        if header == "phealther":
            return f"=H{row_index}"
        if header == "ee ded":
            return "" if ee_deduction <= 0 else f"={min_contribution}-H{row_index}"
        if header == "er ded":
            return "" if er_deduction <= 0 else f"={min_contribution}-I{row_index}"
        if header == "total ded":
            total = ee_deduction + er_deduction
            return "" if total <= 0 else f"=J{row_index}+K{row_index}"
        if header == "phealthrem":
            return f"=H{row_index}+I{row_index}+N(L{row_index})"
        if header == "Monthly Salary":
            return f"=M{row_index}/{phealth_percentage}"
        return src_row.get(header, "")

    if sheet_name == "HDMF":
        if header == "pagibigno":
            return padded_number(src_row.get("pagibigno", ""), 12)
        if header == "ee":
            return src_row.get("pagibig", 0)
        if header == "er":
            return f"=IF(F{row_index}>200,200,F{row_index})"
        return src_row.get(header, "")

    if sheet_name == "SSS":
        sss_value = numeric(src_row.get("sss", 0))
        is_low = sss_value < 250

        if header == "sssno":
            return padded_number(src_row.get("sssno", ""), 10)
        if header == "Monthly Salary":
            return f"=K{row_index}/0.15"
        if header == "sss":
            return sss_value
        if header == "ee":
            return src_row.get("eeshare", 0)
        if header == "ee ded":
            return f"=250-G{row_index}" if is_low else ""
        if header == "ershare":
            return 500 if is_low else src_row.get("ershare", 0)
        if header == "sssrem":
            return 750 if is_low else src_row.get("sssrem", 0)
        if header == "er":
            return 10 if is_low else src_row.get("er", 0)
        if header == "totalsssrem":
            return f"=K{row_index}+L{row_index}"
        return src_row.get(header, "")

    if sheet_name in ["HDMF_LOANS", "SSS_LOANS"]:
        if header in ["pagibigno", "sssno"]:
            width = 12 if header == "pagibigno" else 10
            return padded_number(src_row.get(header, ""), width)
        if header in ["SL Deduction", "CL Deduction"]:
            return ""
        if header == "Total Deduction":
            sl_col = "H" if sheet_name == "HDMF_LOANS" else "G"
            cl_col = "I" if sheet_name == "HDMF_LOANS" else "H"
            return f"=SUM({sl_col}{row_index}:{cl_col}{row_index})"
        if header == "Over Amount Ded":
            src_letter = "G" if sheet_name == "HDMF_LOANS" else "F"
            total_letter = "J" if sheet_name == "HDMF_LOANS" else "I"
            return f"={src_letter}{row_index}-{total_letter}{row_index}"
        return src_row.get(header, "")

    return src_row.get(header, "")


def apply_row_styles(
    sheet_name,
    src_row,
    headers,
    current_row,
    font_regular,
    font_red,
    green_fill,
    border_all,
):
    apply_green = False
    if sheet_name == "HDMF":
        apply_green = numeric(src_row.get("pagibig", 0)) > 200
    elif sheet_name == "SSS":
        apply_green = numeric(src_row.get("sss", 0)) >= 1000

    for column_index, cell in enumerate(current_row, 1):
        column_header = headers[column_index - 1].lower()
        cell.font = font_regular
        if apply_green:
            cell.fill = green_fill
        if column_header in ["ee ded", "er ded", "total ded"]:
            cell.font = font_red
        if any(keyword in column_header for keyword in ["no", "id"]):
            cell.number_format = "@"
        elif "date" in column_header:
            cell.number_format = "mm/dd/yyyy"
        elif isinstance(cell.value, (int, float)) or (
            isinstance(cell.value, str) and cell.value.startswith("=")
        ):
            cell.number_format = "#,##0.00"
        cell.border = border_all


def append_totals(
    worksheet,
    sheet_name,
    headers,
    font_bold,
    border_total_label,
    border_total_value,
):
    worksheet.freeze_panes = "A2"

    last_data_row = worksheet.max_row
    total_row_index = last_data_row + 3
    total_label_cell = worksheet.cell(row=total_row_index, column=1, value="Total")
    total_label_cell.font = font_bold
    total_label_cell.border = border_total_label

    for column_index in range(2, len(headers) + 1):
        worksheet.cell(row=total_row_index, column=column_index).border = (
            border_total_label
        )

    for column_index, _ in enumerate(headers, 1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        should_sum = False
        if sheet_name == "PHIC" and column_index in [8, 9, 10, 11, 12, 13]:
            should_sum = True
        elif sheet_name == "HDMF" and column_index in [6, 7]:
            should_sum = True
        elif sheet_name == "SSS" and column_index in [5, 7, 8, 9, 10, 11, 12]:
            should_sum = True
        elif sheet_name == "HDMF_LOANS" and column_index in range(7, 12):
            should_sum = True
        elif sheet_name == "SSS_LOANS" and column_index in range(6, 11):
            should_sum = True

        if should_sum:
            total_cell = worksheet.cell(row=total_row_index, column=column_index)
            total_cell.value = f"=SUM({column_letter}2:{column_letter}{last_data_row})"
            total_cell.font = font_bold
            total_cell.number_format = "#,##0.00"
            total_cell.border = border_total_value

    if sheet_name == "HDMF":
        grand_total_label = worksheet.cell(
            row=total_row_index + 1,
            column=5,
            value="GRAND TOTAL",
        )
        grand_total_label.font = font_bold
        grand_total_label.border = border_total_label

        grand_total_value = worksheet.cell(row=total_row_index + 1, column=7)
        grand_total_value.value = f"=F{total_row_index}+G{total_row_index}"
        grand_total_value.font = font_bold
        grand_total_value.number_format = "#,##0.00"
        grand_total_value.border = border_total_value


def autofit_columns(worksheet):
    for column in worksheet.columns:
        header_value = column[0].value
        max_length = len(str(header_value)) if header_value else 10
        header_lower = str(header_value).lower() if header_value else ""

        for cell in column:
            if cell.value is None:
                continue

            value = str(cell.value)
            max_length = max(max_length, 15 if value.startswith("=") else len(value))

        column_letter = column[0].column_letter
        if any(
            keyword in header_lower
            for keyword in ["name", "position", "salary", "date", "rate", "rem"]
        ):
            worksheet.column_dimensions[column_letter].width = max_length + 4
        else:
            worksheet.column_dimensions[column_letter].width = max_length + 2


def numeric(value):
    try:
        return float(value) if pd.notna(value) else 0.0
    except (TypeError, ValueError):
        return 0.0


def padded_number(value, width):
    if pd.isna(value) or str(value).strip() == "":
        return ""

    try:
        text = str(int(float(value)))
    except (TypeError, ValueError):
        text = str(value).strip()

    return text.zfill(width) if text.isdigit() else text


if __name__ == "__main__":
    selected_files = select_input()
    if selected_files:
        process_excel_files(selected_files)
        print("\nAll operations complete successfully!")
    else:
        print("\nNo files were selected.")
